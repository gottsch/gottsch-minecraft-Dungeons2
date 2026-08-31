#!/usr/bin/env python3
"""Build an Excel model of a processor list's decay rates, with LIVE formulas.

    python tools/weathering_rates.py src/main/resources/data/dungeons2/worldgen/processor_list/*.json -o out.xlsx

Every authored probability lands in its own blue cell; everything else is a formula off those
cells. So changing one number recalculates the whole block downstream -- the later stages of its
own chain, the shares of every rule below it on the same block, and the totals.

WHAT IT MODELS
--------------
The two compositions that make these files hard to read by eye:

  * A CHAIN (`output_blocks`, or a stage's own entry). Stage k only rolls if k-1 hit, and it
    REPLACES what k-1 produced, so what you see in the world is `reached x (1 - next stage's p)`.
  * A FORK (several rules on one source block). They are alternatives tried in authored order, so
    rule k is only offered what rules 1..k-1 declined. That is the `reach` column.

`agings` is a live cell too: raise or lower it and the stages past the cap drop to zero, which is
the truncation that is invisible in the JSON.

WHAT IT DOES NOT MODEL
----------------------
Chaining ACROSS processor entries -- the thing `classic_entrance_weathering.json` does, where one
`minecraft:rule` entry's output feeds the next entry's input. Each entry gets its own sheet and its
own arithmetic; the compounding between them is not in the workbook. See the notes sheet.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from format_processor_list import strip_comments  # noqa: E402

FONT = "Arial"
INPUT_BLUE = Font(name=FONT, color="0000FF")
FORMULA = Font(name=FONT)
HEADER = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="404040")
BLOCK_FONT = Font(name=FONT, bold=True)
BLOCK_FILL = PatternFill("solid", fgColor="D9D9D9")
NOTE = Font(name=FONT, italic=True, color="808080")
PCT = "0.00%"

COLUMNS = [
    ("Source block", 42),
    ("Rule", 6),
    ("Stage", 6),
    ("Becomes", 42),
    ("Authored p", 12),
    ("Reach", 10),
    ("Reached", 10),
    ("Next p", 9),
    ("SURVIVES", 11),
    ("", 3),
    ("Target", 10),
    ("Author this", 12),
]


def rules_of(processor: dict) -> list[dict]:
    """Normalise both processor shapes to: source block, list of (output, probability)."""
    kind = processor["processor_type"]
    out = []
    if kind in ("dungeons2:aging", "dungeons2:surface_aging"):
        for rule in processor.get("rules", []):
            stages = [(s["block"], s["probability"]) for s in rule["output_blocks"]]
            out.append({"block": rule["block"], "stages": stages,
                        "surface": rule.get("surface", "")})
    elif kind == "minecraft:rule":
        for rule in processor.get("rules", []):
            predicate = rule["input_predicate"]
            block = predicate.get("block", "?")
            if ":" not in block:
                block = "minecraft:" + block
            out.append({"block": block,
                        "stages": [(rule["output_state"]["Name"],
                                    predicate.get("probability", 1.0))],
                        "surface": ""})
    return out


def sheet_for(workbook: Workbook, title: str, processor: dict, source_name: str):
    sheet = workbook.create_sheet(title[:31])
    rules = rules_of(processor)
    if not rules:
        return

    sheet["A1"] = processor["processor_type"]
    sheet["A1"].font = BLOCK_FONT
    sheet["A2"] = f"from {source_name}"
    sheet["A2"].font = NOTE

    # `agings` as a live cell: the truncation it causes is the one thing the JSON hides.
    sheet["E1"] = "agings"
    sheet["E1"].font = FORMULA
    sheet["F1"] = processor.get("agings", 99)
    sheet["F1"].font = INPUT_BLUE
    sheet["G1"] = "<- a stage numbered above this never fires"
    sheet["G1"].font = NOTE

    header_row = 4
    for index, (name, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=name)
        cell.font = HEADER
        if name:
            cell.fill = HEADER_FILL
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Group by source block, preserving authored order -- fork position is the whole point.
    grouped: dict[str, list[dict]] = {}
    for rule in rules:
        grouped.setdefault(rule["block"], []).append(rule)

    row = header_row + 1
    for block, block_rules in grouped.items():
        first_row_of_block = row
        stage_one_rows = []
        survive_rows = []

        for rule_index, rule in enumerate(block_rules, start=1):
            for stage_index, (output, probability) in enumerate(rule["stages"], start=1):
                sheet.cell(row=row, column=1,
                           value=block + (f"  [{rule['surface']}]" if rule["surface"] else "")
                           if stage_index == 1 and rule_index == 1 else "").font = FORMULA
                sheet.cell(row=row, column=2, value=rule_index).font = FORMULA
                sheet.cell(row=row, column=3, value=stage_index).font = FORMULA
                sheet.cell(row=row, column=4, value=output).font = FORMULA

                authored = sheet.cell(row=row, column=5, value=probability)
                authored.font = INPUT_BLUE
                authored.number_format = "0.0000"

                # REACH -- what fraction of the block is still unclaimed when this cell rolls.
                if stage_index > 1:
                    reach = f"=G{row - 1}"                      # the stage above me was reached
                elif rule_index == 1:
                    reach = "=1"                                # nothing has had a turn yet
                else:
                    prior = stage_one_rows[-1]
                    reach = f"=F{prior}*(1-E{prior})"           # the fork: what rule k-1 declined
                sheet.cell(row=row, column=6, value=reach).font = FORMULA
                sheet.cell(row=row, column=6).number_format = PCT

                # REACHED -- and the agings cap, live off F1.
                sheet.cell(row=row, column=7,
                           value=f"=IF(C{row}>$F$1,0,F{row}*E{row})").font = FORMULA
                sheet.cell(row=row, column=7).number_format = PCT

                # NEXT P -- the stage that would replace this one, 0 if this is the last.
                has_next = stage_index < len(rule["stages"])
                next_p = f"=IF(C{row + 1}>$F$1,0,E{row + 1})" if has_next else "=0"
                sheet.cell(row=row, column=8, value=next_p).font = FORMULA
                sheet.cell(row=row, column=8).number_format = "0.0000"

                # SURVIVES -- what a player actually sees, net of the next stage firing instead.
                survives = sheet.cell(row=row, column=9, value=f"=G{row}*(1-H{row})")
                survives.font = FORMULA
                survives.number_format = PCT
                survive_rows.append(row)

                # The fork formula, live: type the share you want, read the number to author.
                if stage_index == 1:
                    sheet.cell(row=row, column=11).font = INPUT_BLUE
                    sheet.cell(row=row, column=11).number_format = PCT
                    sheet.cell(row=row, column=12,
                               value=f'=IF(K{row}="","",K{row}/F{row})').font = FORMULA
                    sheet.cell(row=row, column=12).number_format = "0.0000"
                    stage_one_rows.append(row)

                row += 1

        total = sheet.cell(row=row, column=4, value="changed / unchanged")
        total.font = BLOCK_FONT
        total.alignment = Alignment(horizontal="right")
        changed = "+".join(f"I{r}" for r in survive_rows)
        sheet.cell(row=row, column=9, value=f"={changed}").font = BLOCK_FONT
        sheet.cell(row=row, column=9).number_format = PCT
        sheet.cell(row=row, column=8, value=f"=1-({changed})").font = BLOCK_FONT
        sheet.cell(row=row, column=8).number_format = PCT
        for column in range(1, 10):
            sheet.cell(row=row, column=column).fill = BLOCK_FILL
        row += 2

        _ = first_row_of_block

    sheet.freeze_panes = f"A{header_row + 1}"


def notes_sheet(workbook: Workbook, files: list[str]):
    sheet = workbook.create_sheet("Read me", 0)
    sheet.column_dimensions["A"].width = 110
    lines = [
        ("Dungeons2 weathering rates", BLOCK_FONT),
        ("", FORMULA),
        ("BLUE cells are the authored numbers -- the ones that exist in the JSON. Edit those.", FORMULA),
        ("Everything black is a formula. Change a blue cell and the sheet recomputes downstream.", FORMULA),
        ("", FORMULA),
        ("Reach     what fraction of the source block is still unclaimed when this rule rolls.", FORMULA),
        ("          Rule 1 gets 100%; rule 2 only gets what rule 1 declined. That is the FORK.", FORMULA),
        ("Reached   probability the block gets at least this deep = Reach x Authored p.", FORMULA),
        ("Next p    the stage that would replace this one (0 if it is the last).", FORMULA),
        ("SURVIVES  what a player actually sees = Reached x (1 - Next p). This is the column", FORMULA),
        ("          to compare against the game.", FORMULA),
        ("", FORMULA),
        ("TARGET / AUTHOR THIS -- the fork formula, live. Type the share you want a rule to take", FORMULA),
        ("into Target and 'Author this' gives the probability to put in the JSON. It is", FORMULA),
        ("target / reach, which is why only the first rule on a block can be authored at its", FORMULA),
        ("face value.", FORMULA),
        ("", FORMULA),
        ("agings (cell F1 of each sheet) is live. Lower it and stages past the cap fall to zero --", FORMULA),
        ("that truncation is invisible in the JSON and is a real bug that has shipped.", FORMULA),
        ("", FORMULA),
        ("WHAT THIS DOES NOT MODEL", BLOCK_FONT),
        ("Chaining ACROSS processor entries. classic_entrance_weathering.json runs four", FORMULA),
        ("minecraft:rule entries in a row and entry 2 sees entry 1's output, so a block can reach", FORMULA),
        ("gravel by more than one path and the totals compound between sheets. Each sheet here is", FORMULA),
        ("correct for its own entry; the end-to-end numbers need the whole pipeline simulated.", FORMULA),
        ("", FORMULA),
        ("Generated by tools/weathering_rates.py from:", NOTE),
    ]
    for index, (text, font) in enumerate(lines, start=1):
        cell = sheet.cell(row=index, column=1, value=text)
        cell.font = font
    for offset, name in enumerate(files):
        cell = sheet.cell(row=len(lines) + 1 + offset, column=1, value="    " + name)
        cell.font = NOTE


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("files", nargs="+", type=Path)
    parser.add_argument("-o", "--out", type=Path, required=True)
    args = parser.parse_args()

    workbook = Workbook()
    workbook.remove(workbook.active)

    for path in args.files:
        data = json.loads(strip_comments(path.read_text(encoding="utf-8")))
        short = path.stem.replace("classic_", "").replace("_weathering", "")
        counter = 0
        for processor in data["processors"]:
            if not processor.get("rules"):
                continue
            counter += 1
            kind = processor["processor_type"].split(":")[-1].replace("_", " ")
            sheet_for(workbook, f"{short} {counter} {kind}", processor, path.name)

    notes_sheet(workbook, [str(f) for f in args.files])
    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)
    print(f"wrote {args.out} ({len(workbook.sheetnames)} sheets)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
